import openpyxl
import binary_search_pn_from_price
BOM = openpyxl.load_workbook('上課檔案/Day10/result.xlsx')#開啟BOM file

BOM_sheet = BOM.worksheets[0]#選擇sheet index

# for bom_row in range(2, BOM_sheet.max_row+1): #搜尋每個row的pn
for bom_row in range(2, 100):
    pn = BOM_sheet.cell(column=3, row=bom_row).value #找pn的value
    revision = BOM_sheet.cell(column=4, row=bom_row).value #獲取revision
    if int(revision) < 10: revision = '0'+str(revision) #檢查revision < 10? -> +'0'
    pn = pn+'-'+str(revision) #合併 pn & revision
    print(pn)#顯示pn
    print(binary_search_pn_from_price.binary_search_pn_from_price(pn))#在price_sheet中搜尋pn