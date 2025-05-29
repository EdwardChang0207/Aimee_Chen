import openpyxl
import time
import math
def binary_search_pn_from_price(pn): #從price_sheet用pn當作index -> MFG, price, date
    price = openpyxl.load_workbook('上課檔案/Day10/price.xlsx') #開啟price.xlsx file
    price_sheet = price.worksheets[1]#選擇sheet index
    for i in range(price_sheet.max_row, 0, -1):#排除空的row
        if price_sheet.cell(column=1, row=i).value: break #如果找到第一個“非空”的row，就停止
    lower, upper = i, 2 #設定上下界
    pn = list(map(int, pn.split('-')))#['XXX','XXXXX','XX'] -> pn的前處理
    section = 0 #正搜尋到pn的哪一個區段

    while True: #搜尋直到found or not found
        mid = math.ceil((lower+upper)/2) #計算中間值
        print(f'upper:{upper}, lower:{lower}, mid:{mid}\n')#log

        price_pn = price_sheet.cell(column=1, row=mid).value #目前正在檢查的price_sheet中的pn
        if not(price_pn): #price_pn 不存在 -> del row
            price_sheet.delete_rows(mid, mid)
            continue
        if len(price_pn) != 12: #price_pn 格式錯誤 -> del row
            price_sheet.delete_rows(mid, mid)
            continue
        price_pn = list(map(int, price_pn.split('-')))#['XXX','XXXXX','XX'] -> price_pn前處理

        if price_pn == pn:#找到
            MFG = price_sheet.cell(column=2, row=mid).value
            P = price_sheet.cell(column=3, row=mid).value
            date = price_sheet.cell(column=4, row=mid).value
            return [MFG, P, date] #輸出結果
        
        if price_pn[section] > pn[section]:
            upper = mid
        if price_pn[section] < pn[section]:
            lower = mid
        if price_pn[section] == pn[section]:
            section += 1
        if upper == lower: 
            return
        time.sleep(0.5)