import openpyxl
def load_init_file(file_path):
    target = openpyxl.load_workbook('上課檔案/Day9/result.xlsx')#新檔案
    data = openpyxl.load_workbook(file_path)#來源
    file_name = file_path.split('/')[-1]
    ME = file_name.split('-')[0:3] #['300','00433']
    sheet = data.worksheets[0]#來源的sheet
    target_sheet = target.worksheets[0]#目標的sheet
    row_start = target_sheet.max_row
    for row in range(2, sheet.max_row):
        target_sheet.cell(row = row_start+row-1, column=1).value = f'{ME[0]}-{ME[1]}-{ME[2]}'
        target_sheet_index = 2#目標寫入的col
        for col in range(1,12):#掃描col1~col12
            if col > 4 and col < 8: continue #跳過col5~7
            val = sheet.cell(row=row, column=col).value #讀取data val
            target_sheet.cell(row=row_start+row-1, column=target_sheet_index).value = val #將data val寫入到新檔案
            target_sheet_index += 1 #記錄新檔案目前的col
    target.save('上課檔案/Day9/result.xlsx')#存檔
