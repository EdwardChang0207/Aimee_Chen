import openpyxl

def title_init(path):
    target = openpyxl.load_workbook('上課檔案/Day9/result.xlsx')#新檔案
    data = openpyxl.load_workbook(path)#來源

    sheet = data.worksheets[0]#來源的sheet
    target_sheet = target.worksheets[0]#目標的sheet
    
    target_sheet_index = 1#目標寫入的col
    for col in range(1,12):#掃描col1~col12
        if col > 4 and col < 8: continue #跳過col5~7
        val = sheet.cell(row=1, column=col).value #讀取data val
        target_sheet.cell(row=1, column=target_sheet_index).value = val #將data val寫入到新檔案
        target_sheet_index += 1 #記錄新檔案目前的col
    target_sheet.insert_cols(1)
    target_sheet.cell(column=1, row=1).value = 'ME'
    target.save('上課檔案/Day9/result.xlsx')#存檔
    