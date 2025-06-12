import openpyxl

def del_empty(file_name):

    workbook = openpyxl.load_workbook(file_name)

    #刪除多餘的sheet
    names = workbook.sheetnames #全部sheet的名稱
    for i in range(len(names)-1, 0, -1):
        workbook.remove(workbook[names[i]])

    #前處理過後的檔案
    workbook.create_sheet('del_empty')

    sheet = workbook.worksheets[0]
    result = workbook.worksheets[1]

    r = 1#寫入result 的第r個row
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=12).value:
            for j in range(1,21):
                result.cell(row=r, column=j).value = sheet.cell(row=i, column=j).value
            r += 1
    workbook.save(file_name)
