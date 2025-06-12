import openpyxl
def sku_sum(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.worksheets[2]
    workbook.create_sheet('sku_sum')
    result = workbook.worksheets[3]

    current = sheet.cell(column=1, row=1).value #現在的機種
    result.cell(column=1, row=1).value = current
    module_no = 1
    k = 0
    for i in range(1,sheet.max_row+1):
        module = sheet.cell(column=1, row=i).value #目前掃描的機種
        if current != module: #換機種
            current = module
            result.cell(column=2,row=module_no).value = float(f'{k:.2f}')
            module_no += 1
            result.cell(column=1,row=module_no).value = module
            k = 0
        x = str(sheet.cell(column=2, row=i).value)
        k += float(x)
    result.cell(column=2,row=module_no).value = float(f'{k:.2f}') #f-string -> str
    workbook.save(file_name)