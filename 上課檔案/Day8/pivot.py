import openpyxl
def pivot(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.worksheets[1]
    workbook.create_sheet('result')
    result = workbook.worksheets[2]
    current = sheet.cell(column=1, row=2).value #現在的機種
    result.cell(column=1, row=1).value = current
    module_no = 1
    k = 0
    current_MFG = sheet.cell(column=12, row=2).value
    result.cell(column=3, row=1).value = current_MFG
    for i in range(2,sheet.max_row+1):
        module = sheet.cell(column=1, row=i).value #目前掃描的機種
        MFG = sheet.cell(column=12, row=i).value
        if not(MFG): continue
        if current != module or current_MFG != MFG: #換機種
            
            current = module
            result.cell(column=2,row=module_no).value = float(f'{k:.2f}')
            module_no += 1
            result.cell(column=1,row=module_no).value = module
            current_MFG = MFG
            result.cell(column=3,row=module_no).value = MFG
            k = 0
        x = sheet.cell(column=11, row=i).value
        if not x: x = 0
        k += x 
    result.cell(column=2,row=module_no).value = float(f'{k:.2f}') #f-string -> str
    workbook.save(file_name)