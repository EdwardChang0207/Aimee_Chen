import openpyxl

wk = openpyxl.load_workbook('上課檔案/Day16/BOM_download.xlsx')
sh = wk.worksheets[0]
sh.cell(row=2, column=1).value = '300-00184-01'
sh.cell(row=3, column=1).value = '300-02407-01'

wk.save('上課檔案/Day16/BOM_download.xlsx')
wk.close()