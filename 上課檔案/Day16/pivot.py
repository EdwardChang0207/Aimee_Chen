import openpyxl
#get source file
source = openpyxl.load_workbook('上課檔案/Day16/result.xlsx')
source_sheet = source.worksheets[0]
#select target sheet
if len(source.worksheets) < 2:
    source.create_sheet('pivot')
target_sheet = source.worksheets[1]

PN = set()
for i in range(2, source_sheet.max_row+1):
    item_number = source_sheet.cell(row=i, column=4).value
    PN.add(item_number)
PN_sku = dict()
#PN_sku -> sku -> [quantity, mfg]
for i in PN:
    PN_sku[i] = dict()
for i in range(2, source_sheet.max_row+1):
    item_number = source_sheet.cell(row=i, column=4).value
    sku = source_sheet.cell(row=i, column=1).value
    mfg = source_sheet.cell(row=i, column=17).value
    PN_sku[item_number][sku]=[0,mfg]

for i in range(2, source_sheet.max_row+1):
    item_number = source_sheet.cell(row=i, column=4).value
    sku = source_sheet.cell(row=i, column=1).value
    quantity = source_sheet.cell(row=i, column=10).value
    PN_sku[item_number][sku][0]+=quantity

print(PN_sku)
cur_row = 1
PN = list(PN)
print(PN)
for i in range(len(PN)):
    for sku in PN_sku[PN[i]].keys():
        target_sheet.cell(row=cur_row, column=1).value = PN[i] #PN
        target_sheet.cell(row=cur_row, column=2).value = sku#sku
        target_sheet.cell(row=cur_row, column=3).value = PN_sku[PN[i]][sku][0]#Q
        target_sheet.cell(row=cur_row, column=6).value = PN_sku[PN[i]][sku][1]#MFG
        cur_row += 1

sku_demand_dict = dict()
sku_demand = openpyxl.load_workbook('上課檔案/Day16/BOM_download.xlsx')
sku_demand_sheet = sku_demand.worksheets[0]
for i in range(1, sku_demand_sheet.max_row+1):
    sku = sku_demand_sheet.cell(row=i, column=2).value
    demand = sku_demand_sheet.cell(row=i, column=3).value
    sku_demand_dict[sku] = demand

for i in range(1, target_sheet.max_row+1):
    sku = target_sheet.cell(row=i, column=2).value
    demand = sku_demand_dict[sku]
    quantity = target_sheet.cell(row=i, column=3).value
    target_sheet.cell(row=i, column=4).value = demand
    target_sheet.cell(row=i, column=5).value = demand*quantity

source.save('上課檔案/Day16/result.xlsx')