import openpyxl
# openpyxl.load_workbook("Python D12\Download BOM.py")
def Run_ME_BOM(path):
    load_bom = openpyxl.load_workbook(path) #把 Excel 檔案的路徑（字串）傳進來
    ME_BOM = load_bom.worksheets[0]

# 把第一欄的所有資料讀成一個 Python list
    ME_BOM_List = []                       #建立一個空的 list，準備把 Excel 裡的資料一筆一筆放進來
    for i in range(1, ME_BOM.max_row +1 ):
        ME = ME_BOM.cell(column=1, row=i).value
        ME_BOM_List.append(ME)             #把讀到的這筆資料加到 ME_BOM_List 中每跑一圈就會多一筆資料
    return ME_BOM_List                     #把整個 list（第一欄資料）回傳出去