import openpyxl
#get source file
source = openpyxl.load_workbook('上課檔案/Day16/result.xlsx')
source_sheet = source.worksheets[0]
#select target sheet
if len(source.worksheets) < 3:
    source.create_sheet('MFG')
target_sheet = source.worksheets[2]

MFG = set()
for i in range(2, source_sheet.max_row+1):
    item_number = source_sheet.cell(row=i, column=17).value
    MFG.add(item_number)
MFG_SKU = dict()
#mfg -> sku -> PN -> total_demand
for i in MFG:
    MFG_SKU[i] = dict()
for i in range(2, source_sheet.max_row+1):
    mfg = source_sheet.cell(row=i, column=17).value
    sku = source_sheet.cell(row=i, column=1).value
    MFG_SKU[mfg][sku] = dict()

for i in range(2, source_sheet.max_row+1):
    mfg = source_sheet.cell(row=i, column=17).value
    sku = source_sheet.cell(row=i, column=1).value
    item_number = source_sheet.cell(row=i, column=4).value
    MFG_SKU[mfg][sku][item_number] = 0

for i in range(2, source_sheet.max_row+1):
    mfg = source_sheet.cell(row=i, column=17).value
    sku = source_sheet.cell(row=i, column=1).value
    item_number = source_sheet.cell(row=i, column=4).value
    quantity = source_sheet.cell(row=i, column=10).value
    MFG_SKU[mfg][sku][item_number]+=quantity

# print(MFG_SKU)

cur_row = 1
MFG = list(MFG)
print(MFG)
for i in range(len(MFG)):
    for sku in MFG_SKU[MFG[i]].keys():
        for pn in MFG_SKU[MFG[i]][sku].keys():
            target_sheet.cell(row=cur_row, column=1).value = MFG[i] #MFG
            target_sheet.cell(row=cur_row, column=2).value = sku#SKU
            target_sheet.cell(row=cur_row, column=3).value = pn#PN
            target_sheet.cell(row=cur_row, column=4).value = MFG_SKU[MFG[i]][sku][pn]#Q
            cur_row += 1

# sku_demand_dict = dict()
# sku_demand = openpyxl.load_workbook('上課檔案/Day16/BOM_download.xlsx')
# sku_demand_sheet = sku_demand.worksheets[0]
# for i in range(1, sku_demand_sheet.max_row+1):
#     sku = sku_demand_sheet.cell(row=i, column=2).value
#     demand = sku_demand_sheet.cell(row=i, column=3).value
#     sku_demand_dict[sku] = demand

# for i in range(1, target_sheet.max_row+1):
#     sku = target_sheet.cell(row=i, column=2).value
#     demand = sku_demand_dict[sku]
#     quantity = target_sheet.cell(row=i, column=3).value
#     target_sheet.cell(row=i, column=4).value = demand
#     target_sheet.cell(row=i, column=5).value = demand*quantity

source.save('上課檔案/Day16/result.xlsx')