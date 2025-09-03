import glob
import openpyxl
def combine_BOM(path):
    #create new wk, ws
    ME_BOM_ALL = openpyxl.Workbook()
    ME_BOM_ALL_SW = ME_BOM_ALL.worksheets[0]
    
    #create sku-name searching table
    source = openpyxl.load_workbook('上課檔案/Day16/BOM_download.xlsx')
    source_sheet = source.worksheets[0]
    sku_name_searching_table = dict()
    sku_demand_searching_table = dict()
    for row in range(1, source_sheet.max_row+1):
        k = source_sheet.cell(row=row, column=1).value
        if not k:break
        k = k.split('-')
        k = f'{k[0]}-{k[1]}'
        sku_name = source_sheet.cell(row=row, column=2).value
        sku_name_searching_table[k]=sku_name
        sku_demand = source_sheet.cell(row=row, column=3).value
        sku_demand_searching_table[k]=sku_demand
    #search all xlsx files
    pattern = f'{path}/*.xlsx'
    matching_files = glob.glob(pattern)

    #判斷title是否已經存入
    title = False

    #對每一個檔案進行處理
    for f in matching_files:
        #抓取filename
        file_name = f.split('/')[-1]
        file_name = file_name.split('-')[0:3]
        file_name[2] = file_name[2][1::]
        #判斷revision是否要補0
        if len(file_name[2]) == 1: file_name[2] = '0'+ file_name[2]
        #合併BOM_NO
        BOM_NO = f'{file_name[0]}-{file_name[1]}-{file_name[2]}'
        BOM_NO_KEY = f'{file_name[0]}-{file_name[1]}'
        #載入來源檔
        wk = openpyxl.load_workbook(f)
        ws = wk.worksheets[0]

        titles = []#存取row1的標題
        for col in range(1, ws.max_column+1):
            titles.append(ws.cell(row=1, column=col).value)
        #判斷vendor是否存在 if yes -> del
        if 'Vendor 1' in titles:
            start = titles.index('Vendor 1')
            end = titles.index('Manufacturer 1')
            count = end-start
            ws.delete_cols(start+1, count)
        #判斷 sub rank是否存在 if -> del 0, col
        if 'Sub. rank' in titles:
            rm_row = []
            for r in range(1, ws.max_row+1):
                if ws.cell(row=r, column=2).value == 0:
                    rm_row.append(r)
            rm_row.reverse()
            for r in rm_row:
                ws.delete_rows(r,1)
            ws.delete_cols(2,1)
        titles = []#存取row1的標題
        for col in range(1, ws.max_column+1):
            titles.append(ws.cell(row=1, column=col).value)
        mf1 = titles.index('Manufacturer 1')
        ws.delete_cols(mf1+2, ws.max_column-mf1)

        #記錄輸出檔案目前寫入到哪一個row
        target_file_write_start_point = ME_BOM_ALL_SW.max_row

        #寫入A,B col的標題
        if not title: 
            t = 1 #是第一個檔案 -> 從row 1開始寫入
            ME_BOM_ALL_SW.cell(row=1, column=1).value = 'sku_name'
            ME_BOM_ALL_SW.cell(row=1, column=1+1).value = 'BOM_version'
            ME_BOM_ALL_SW.cell(row=1, column=2+1).value = 'pn_version'
            

        else: t = 2 #不是第一個檔案 -> 從row 2開始寫入

        #讀取來源檔案的每一個cell
        for r in range(t, ws.max_row+1): #loop through every row
            if ws.cell(row=r,column=1).value == 'Total': break #if 到了total那一個row -> stop

            if r > 1:#不是標題的row
                #寫入BOM_NO
                ME_BOM_ALL_SW.cell(row=target_file_write_start_point+r-1, column=1).value = sku_name_searching_table[BOM_NO_KEY]
                ME_BOM_ALL_SW.cell(row=target_file_write_start_point+r-1, column=1+1).value = BOM_NO
                ME_BOM_ALL_SW.cell(row=target_file_write_start_point+r-1, column=sku_demand_col).value = sku_demand_searching_table[BOM_NO_KEY]

                #寫入PN
                item_number = ws.cell(row=r, column=2+1).value
                revision = str(ws.cell(row=r, column=3+1).value)
                #revision < 10 -> 補0
                if len(revision) == 1: revision = '0'+revision
                ME_BOM_ALL_SW.cell(row=target_file_write_start_point+r-1, column=2+1).value = f'{item_number}-{revision}'
            
            for c in range(1, ws.max_column+1):#loop through every col
                i = ws.cell(row=r, column=c).value
                target_cell = ME_BOM_ALL_SW.cell(row=target_file_write_start_point+r-1, column=c+2)
                target_cell.value = i
            
            if t == 1:
                print(ME_BOM_ALL_SW.max_column)
                ME_BOM_ALL_SW.insert_cols(ME_BOM_ALL_SW.max_column+1)
                print(ME_BOM_ALL_SW.max_column)
                ME_BOM_ALL_SW.cell(row=1, column=ME_BOM_ALL_SW.max_column).value = 'demand'
                sku_demand_col = ME_BOM_ALL_SW.max_column

        ME_BOM_ALL.save('上課檔案/Day16/result.xlsx')
        title = True

combine_BOM('上課檔案/Day16/BOM_file')
