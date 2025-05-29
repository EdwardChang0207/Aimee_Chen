from openpyxl import Workbook
from random import randint
wb = Workbook()
ws = wb.active
ws.title = 'Test'
ws1 = wb.create_sheet('test')
ws2 = wb['Test']

# ws['A4'] = 4
# d = ws.cell(column=1, row=4, value=5)
for i in range(1,1001):
    col, row = (i//10)+1, i%10
    if row == 0: 
        row = 10
        col -= 1
    ws.cell(column=col,row=row,value=randint(1,100))

target = ws.cell(column=1, row=1).value
count = 0
for i in range(1,1001):
    col = i % 100
    row = i//100 + 1
    if col == 0:
        col = 100
    if ws.cell(column=col, row=row).value == target:
        count += 1
count -= 1
print(count)

wb.save('test.xlsx')