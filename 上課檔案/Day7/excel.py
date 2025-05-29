from openpyxl import Workbook
from random import randint
wb = Workbook()
ws = wb.active
ws.title = 'Test'
ws1 = wb.create_sheet('test')
ws2 = wb['Test']

for i in range(1,11):
    ws.cell(row=i,column=1).value = randint(0,2)
for i in range(1,11):
    ws.cell(row=i,column=2).value = randint(1,10)

s = 0
for i in range(1,11):
    if ws.cell(row=i, column=1).value == 1:
        s += ws.cell(row=i, column=2).value
ws.cell(column=1, row=11).value = s
wb.save('test.xlsx')